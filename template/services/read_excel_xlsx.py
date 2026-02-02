"""
Module đọc file Excel có công thức và giải thích các trường hợp
"""
import pandas as pd
from typing import List, Dict, Any, Union
import openpyxl
from openpyxl.utils import get_column_letter
import logging

logger = logging.getLogger(__name__)

def read_excel_with_formulas(file_path: str, sheet_name: int | str = 0) -> tuple:
    """
    Đọc file Excel và trả về CÁ GIÁ TRỊ và CÔNG THỨC (nếu có)
    
    Returns:
        tuple: (data_values, formulas_info)
            - data_values: Mảng dict với giá trị đã tính toán
            - formulas_info: Dict chứa thông tin về các công thức
    """
    try:
        # 1. Đọc giá trị đã tính toán bằng pandas
        df = pd.read_excel(file_path, sheet_name=sheet_name)
        data_values = df.to_dict(orient='records')
        
        # 2. Đọc công thức bằng openpyxl
        workbook = openpyxl.load_workbook(file_path, data_only=False)
        
        # Lấy sheet theo tên hoặc index
        if isinstance(sheet_name, int):
            sheet = workbook.worksheets[sheet_name]
        else:
            sheet = workbook[sheet_name]
        
        # Thu thập thông tin về công thức
        formulas_info = {
            'has_formulas': False,
            'formula_cells': [],
            'details': {}
        }
        
        # Duyệt qua tất cả các ô có dữ liệu
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):  # Bỏ qua header
            for col_idx, cell in enumerate(row, start=1):
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    formulas_info['has_formulas'] = True
                    cell_address = f"{get_column_letter(col_idx)}{row_idx}"
                    
                    # Lưu thông tin công thức
                    formula_detail = {
                        'cell': cell_address,
                        'formula': cell.value,
                        'row_index': row_idx - 2,  # Index trong data array
                        'column_name': sheet.cell(1, col_idx).value  # Tên cột từ header
                    }
                    
                    formulas_info['formula_cells'].append(cell_address)
                    formulas_info['details'][cell_address] = formula_detail
        
        workbook.close()
        
        return data_values, formulas_info
    
    except Exception as e:
        logger.error(f"Lỗi khi đọc file Excel với công thức: {str(e)}")
        return [], {'has_formulas': False, 'formula_cells': [], 'details': {}}


def get_sheet_names(file_path: str) -> List[str]:
    """
    Lấy danh sách tên các sheet trong file Excel
    
    Args:
        file_path (str): Đường dẫn đến file Excel
    
    Returns:
        List[str]: Danh sách tên các sheet
    
    Example:
        >>> sheets = get_sheet_names('data.xlsx')
        >>> print(sheets)
        ['Sheet1', 'Sheet2', 'Dữ liệu']
    """
    try:
        # Cách 1: Dùng pandas (nhanh hơn)
        xl_file = pd.ExcelFile(file_path)
        return xl_file.sheet_names
        
        # Cách 2: Dùng openpyxl (nếu cần)
        # workbook = openpyxl.load_workbook(file_path, read_only=True)
        # return workbook.sheetnames
    
    except Exception as e:
        logger.error(f"Lỗi khi đọc sheet names: {str(e)}")
        return []


def read_excel_to_array(file_path: str, sheet_name: Union[int, str] = 0) -> List[Dict[str, Any]]:
    """
    Đọc 1 SHEET cụ thể từ file Excel và trả về mảng đối tượng
    
    Args:
        file_path (str): Đường dẫn đến file Excel
        sheet_name (int | str): Tên hoặc index của sheet (mặc định: 0 = sheet đầu tiên)
    
    Returns:
        List[Dict[str, Any]]: Mảng các đối tượng
    
    Example:
        >>> # Đọc sheet đầu tiên
        >>> data = read_excel_to_array('data.xlsx')
        >>> 
        >>> # Đọc sheet theo tên
        >>> data = read_excel_to_array('data.xlsx', 'Nhân viên')
        >>> 
        >>> # Đọc sheet thứ 2 (index = 1)
        >>> data = read_excel_to_array('data.xlsx', 1)
    """
    try:
        df = pd.read_excel(file_path, sheet_name=sheet_name)
        data_array = df.to_dict(orient='records')
        return data_array
    
    except Exception as e:
        logger.error(f"Lỗi khi đọc sheet '{sheet_name}': {str(e)}")
        return []


def read_all_sheets_to_dict(file_path: str) -> Dict[str, List[Dict[str, Any]]]:
    """
    Đọc TẤT CẢ các sheet từ file Excel và trả về dictionary
    
    Returns:
        Dict[str, List[Dict]]: Dictionary với:
            - Key: Tên sheet
            - Value: Mảng đối tượng của sheet đó
    
    Example:
        >>> all_data = read_all_sheets_to_dict('data.xlsx')
        >>> print(all_data.keys())  # dict_keys(['Sheet1', 'Sheet2', 'Nhân viên'])
        >>> 
        >>> # Truy cập dữ liệu của từng sheet
        >>> nhan_vien = all_data['Nhân viên']
        >>> print(nhan_vien[0])
    """
    try:
        # Đọc tất cả sheet cùng lúc
        all_sheets = pd.read_excel(file_path, sheet_name=None)  # None = đọc tất cả
        
        # Chuyển đổi mỗi sheet thành list of dict
        result = {}
        for sheet_name, df in all_sheets.items():
            result[sheet_name] = df.to_dict(orient='records')
        
        return result
    
    except Exception as e:
        logger.error(f"Lỗi khi đọc tất cả sheets: {str(e)}")
        return {}


def read_all_sheets_to_array(file_path: str) -> List[Dict[str, Any]]:
    """
    Đọc TẤT CẢ các sheet và GỘP CHUNG thành 1 mảng duy nhất
    Thêm cột 'sheet_name' để phân biệt dữ liệu từ sheet nào
    
    Returns:
        List[Dict[str, Any]]: Mảng đối tượng gộp từ tất cả các sheet
    
    Example:
        >>> all_data = read_all_sheets_to_array('data.xlsx')
        >>> print(all_data[0])
        {'sheet_name': 'Sheet1', 'Mã NV': 'NV001', ...}
    """
    try:
        all_sheets = pd.read_excel(file_path, sheet_name=None)
        
        combined_data = []
        for sheet_name, df in all_sheets.items():
            # Thêm cột sheet_name
            df['sheet_name'] = sheet_name
            
            # Chuyển đổi và thêm vào mảng tổng
            sheet_data = df.to_dict(orient='records')
            combined_data.extend(sheet_data)
        
        return combined_data
    
    except Exception as e:
        logger.error(f"Lỗi khi đọc và gộp sheets: {str(e)}")
        return []


def read_specific_sheets(file_path: str, sheet_names: List[Union[int, str]]) -> Dict[str, List[Dict[str, Any]]]:
    """
    Đọc NHIỀU sheet cụ thể (không phải tất cả)
    
    Args:
        file_path (str): Đường dẫn file Excel
        sheet_names (List[int | str]): Danh sách tên hoặc index các sheet cần đọc
    
    Returns:
        Dict[str, List[Dict]]: Dictionary chứa dữ liệu các sheet được chọn
    
    Example:
        >>> # Đọc 2 sheet cụ thể
        >>> data = read_specific_sheets('data.xlsx', ['Nhân viên', 'Phòng ban'])
        >>> 
        >>> # Đọc sheet theo index
        >>> data = read_specific_sheets('data.xlsx', [0, 2, 3])
    """
    try:
        result = {}
        
        for sheet_name in sheet_names:
            df = pd.read_excel(file_path, sheet_name=sheet_name)
            
            # Lấy tên sheet thật (nếu input là index)
            if isinstance(sheet_name, int):
                all_names = get_sheet_names(file_path)
                actual_name = all_names[sheet_name]
            else:
                actual_name = sheet_name
            
            result[actual_name] = df.to_dict(orient='records')
        
        return result
    
    except Exception as e:
        logger.error(f"Lỗi khi đọc sheets: {str(e)}")
        return {}


def get_sheet_info(file_path: str) -> Dict[str, Dict[str, Any]]:
    """
    Lấy thông tin tổng quan về tất cả các sheet trong file
    
    Returns:
        Dict với thông tin:
            - sheet_name: {row_count, column_count, columns}
    
    Example:
        >>> info = get_sheet_info('data.xlsx')
        >>> print(info)
        {
            'Sheet1': {'row_count': 100, 'column_count': 5, 'columns': ['A', 'B', ...]},
            'Sheet2': {'row_count': 50, 'column_count': 3, 'columns': ['X', 'Y', 'Z']}
        }
    """
    try:
        all_sheets = pd.read_excel(file_path, sheet_name=None)
        
        info = {}
        for sheet_name, df in all_sheets.items():
            info[sheet_name] = {
                'row_count': len(df),
                'column_count': len(df.columns),
                'columns': list(df.columns)
            }
        
        return info
    
    except Exception as e:
        logger.error(f"Lỗi khi lấy thông tin sheets: {str(e)}")
        return {}


def matching_data(data_cong, data_luong):
    try:
        length_cong = len(data_cong)
        length_luong = len(data_luong)

        if length_cong != length_luong:
            logger.warning(f"Cảnh báo: Số dòng chấm công ({length_cong}) và lương ({length_luong}) không khớp!")
            return []
        
        logger.info("Đang ghép dữ liệu chấm công và lương...")
        combined_data = []
        for i in range(length_cong):
            for j in range(length_luong):
                if data_cong[i]['Mã nhân viên'] == data_luong[j]['Mã nhân viên']:
                    combined_record = {**data_cong[i], **data_luong[j]}
                    combined_data.append(combined_record)
                    break  # Tìm thấy rồi, không cần kiểm tra tiếp

        return combined_data
    
    except Exception as e:
        logger.error(f"Lỗi khi ghép dữ liệu: {str(e)}")
        return []