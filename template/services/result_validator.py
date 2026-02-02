#!/usr/bin/env python3
"""
Result Validator - Kiểm tra kết quả và thông báo lỗi sau khi merge và update công thức

Chức năng:
1. Kiểm tra số dòng tương ứng với số nhân viên
2. Kiểm tra ô trống trong các cột bắt buộc
3. Kiểm tra độ chính xác công thức (formula adjustment)
4. Kiểm tra công thức tham chiếu sheet khác (VLOOKUP, etc.)
5. Báo cáo "thiếu thông tin" khi không tìm được dữ liệu
"""
import openpyxl
import re
import logging
from typing import Dict, List, Any, Tuple, Optional
from openpyxl.cell.cell import Cell
from openpyxl.utils import get_column_letter

# Logging configured in main.py
logger = logging.getLogger(__name__)


class ResultValidator:
    """Validate result file after merging and formula adjustment"""
    
    def __init__(self, file_path: str):
        self.file_path = file_path
        self.wb = openpyxl.load_workbook(file_path, data_only=False)
        self.validation_errors = []
        self.validation_warnings = []
        self.validation_info = []
    
    def validate_all(self, expected_employee_count: int = None) -> Dict[str, Any]:
        """
        Validate toàn bộ file result
        
        Args:
            expected_employee_count: Số nhân viên mong đợi
            
        Returns:
            Dict chứa kết quả validation với các key:
            - success: bool
            - errors: List[Dict] - Các lỗi nghiêm trọng
            - warnings: List[Dict] - Các cảnh báo
            - info: List[Dict] - Thông tin bổ sung
            - summary: Dict - Tóm tắt kết quả
        """
        logger.info(f"=== Starting validation for {self.file_path} ===")
        
        results = {
            "file": self.file_path,
            "success": True,
            "errors": [],
            "warnings": [],
            "info": [],
            "sheets": {}
        }
        
        # 1. Validate main salary sheet (usually first sheet or '1.BL CHI TIẾT')
        main_sheet_name = self._find_main_salary_sheet()
        if main_sheet_name:
            logger.info(f"Found main salary sheet: {main_sheet_name}")
            
            main_sheet_result = self._validate_salary_sheet(
                main_sheet_name, 
                expected_employee_count
            )
            results["sheets"][main_sheet_name] = main_sheet_result
            results["errors"].extend(main_sheet_result.get("errors", []))
            results["warnings"].extend(main_sheet_result.get("warnings", []))
            results["info"].extend(main_sheet_result.get("info", []))
        
        # 2. Validate BCC attendance sheet if exists
        bcc_sheet_name = self._find_bcc_sheet()
        if bcc_sheet_name:
            logger.info(f"Found BCC attendance sheet: {bcc_sheet_name}")
            
            bcc_result = self._validate_bcc_sheet(
                bcc_sheet_name,
                expected_employee_count
            )
            results["sheets"][bcc_sheet_name] = bcc_result
            results["errors"].extend(bcc_result.get("errors", []))
            results["warnings"].extend(bcc_result.get("warnings", []))
            results["info"].extend(bcc_result.get("info", []))
        
        # 3. Validate other sheets
        for sheet_name in self.wb.sheetnames:
            if sheet_name not in [main_sheet_name, bcc_sheet_name]:
                sheet_result = self._validate_generic_sheet(sheet_name)
                if sheet_result:
                    results["sheets"][sheet_name] = sheet_result
                    results["warnings"].extend(sheet_result.get("warnings", []))
        
        # 4. Generate summary
        results["summary"] = {
            "total_errors": len(results["errors"]),
            "total_warnings": len(results["warnings"]),
            "total_info": len(results["info"]),
            "sheets_validated": len(results["sheets"])
        }
        
        # Set success flag
        results["success"] = len(results["errors"]) == 0
        
        logger.info(f"=== Validation completed: {results['summary']} ===")
        
        return results
    
    def _find_main_salary_sheet(self) -> Optional[str]:
        """Tìm sheet lương chính"""
        # Priority order
        priority_names = [
            "1.BL CHI TIẾT", "1. BL CHI TIẾT", "BL CHI TIẾT",
            "BẢNG LƯƠNG", "BANG LUONG", "SALARY"
        ]
        
        for name in priority_names:
            if name in self.wb.sheetnames:
                return name
        
        # Return first sheet as fallback
        return self.wb.sheetnames[0] if self.wb.sheetnames else None
    
    def _find_bcc_sheet(self) -> Optional[str]:
        """Tìm sheet chấm công (BCC)"""
        bcc_keywords = ["BCC", "1.BCC", "CHẤM CÔNG", "CHAM CONG", "ATTENDANCE"]
        
        for sheet_name in self.wb.sheetnames:
            sheet_upper = sheet_name.upper()
            if any(keyword in sheet_upper for keyword in bcc_keywords):
                return sheet_name
        
        return None
    
    def _validate_salary_sheet(
        self, 
        sheet_name: str, 
        expected_employee_count: Optional[int]
    ) -> Dict[str, Any]:
        """
        Validate sheet lương chính
        
        Checks:
        1. Số dòng vs số nhân viên
        2. Các ô trống trong cột bắt buộc
        3. Công thức điều chỉnh đúng
        4. VLOOKUP và tham chiếu sheet khác
        """
        ws = self.wb[sheet_name]
        result = {
            "sheet_name": sheet_name,
            "errors": [],
            "warnings": [],
            "info": []
        }
        
        # Find data range
        header_row, data_start_row, data_end_row = self._find_data_range(ws)
        
        if not data_start_row:
            result["warnings"].append({
                "type": "DATA_RANGE_NOT_FOUND",
                "message": f"Không tìm thấy vùng dữ liệu trong sheet {sheet_name}"
            })
            return result
        
        result["info"].append({
            "type": "DATA_RANGE",
            "message": f"Header: dòng {header_row}, Dữ liệu: dòng {data_start_row} - {data_end_row}",
            "header_row": header_row,
            "data_start_row": data_start_row,
            "data_end_row": data_end_row
        })
        
        actual_row_count = data_end_row - data_start_row + 1
        
        # 1. CHECK: Số dòng vs số nhân viên
        if expected_employee_count:
            if actual_row_count < expected_employee_count:
                result["errors"].append({
                    "type": "ROW_COUNT_MISMATCH",
                    "severity": "HIGH",
                    "message": f"Thiếu dòng dữ liệu: có {actual_row_count} dòng nhưng cần {expected_employee_count} nhân viên",
                    "expected": expected_employee_count,
                    "actual": actual_row_count
                })
            elif actual_row_count > expected_employee_count:
                result["warnings"].append({
                    "type": "EXTRA_ROWS",
                    "message": f"Có {actual_row_count} dòng nhưng chỉ {expected_employee_count} nhân viên",
                    "expected": expected_employee_count,
                    "actual": actual_row_count
                })
            else:
                result["info"].append({
                    "type": "ROW_COUNT_OK",
                    "message": f"✓ Số dòng khớp: {actual_row_count} dòng cho {expected_employee_count} nhân viên"
                })
        
        # 2. CHECK: Các cột bắt buộc và ô trống
        required_columns = self._identify_required_columns(ws, header_row)
        empty_cell_issues = self._check_empty_required_cells(
            ws, 
            required_columns, 
            data_start_row, 
            data_end_row
        )
        result["warnings"].extend(empty_cell_issues)
        
        # 3. CHECK: Formula adjustment
        formula_issues = self._check_formula_adjustment(
            ws, 
            data_start_row, 
            data_end_row
        )
        result["errors"].extend([e for e in formula_issues if e.get("severity") == "HIGH"])
        result["warnings"].extend([w for w in formula_issues if w.get("severity") != "HIGH"])
        
        # 4. CHECK: VLOOKUP và cross-sheet references
        vlookup_issues = self._check_vlookup_formulas(
            ws,
            data_start_row,
            data_end_row
        )
        result["warnings"].extend(vlookup_issues)
        
        return result
    
    def _validate_bcc_sheet(
        self,
        sheet_name: str,
        expected_employee_count: Optional[int]
    ) -> Dict[str, Any]:
        """Validate sheet BCC (chấm công)"""
        ws = self.wb[sheet_name]
        result = {
            "sheet_name": sheet_name,
            "errors": [],
            "warnings": [],
            "info": []
        }
        
        # Find data range
        header_row, data_start_row, data_end_row = self._find_data_range(ws)
        
        if not data_start_row:
            result["warnings"].append({
                "type": "DATA_RANGE_NOT_FOUND",
                "message": f"Không tìm thấy vùng dữ liệu trong sheet {sheet_name}"
            })
            return result
        
        actual_row_count = data_end_row - data_start_row + 1
        
        # CHECK: Row count
        if expected_employee_count and actual_row_count != expected_employee_count:
            result["warnings"].append({
                "type": "BCC_ROW_MISMATCH",
                "message": f"Sheet BCC có {actual_row_count} dòng, mong đợi {expected_employee_count}",
                "expected": expected_employee_count,
                "actual": actual_row_count
            })
        else:
            result["info"].append({
                "type": "BCC_ROW_OK",
                "message": f"✓ Sheet BCC có {actual_row_count} dòng dữ liệu"
            })
        
        # CHECK: Key attendance columns have data
        attendance_columns = ["ngày công", "số ngày", "working days", "công"]
        attendance_issues = self._check_attendance_data(
            ws,
            header_row,
            data_start_row,
            data_end_row,
            attendance_columns
        )
        result["warnings"].extend(attendance_issues)
        
        return result
    
    def _validate_generic_sheet(self, sheet_name: str) -> Optional[Dict[str, Any]]:
        """Validate các sheet khác (không phải lương hoặc BCC)"""
        # For now, just check for #REF! errors
        ws = self.wb[sheet_name]
        result = {
            "sheet_name": sheet_name,
            "warnings": []
        }
        
        ref_errors = self._find_ref_errors(ws)
        if ref_errors:
            result["warnings"].extend(ref_errors)
            return result
        
        return None
    
    def _find_data_range(self, ws) -> Tuple[Optional[int], Optional[int], Optional[int]]:
        """
        Tìm header row, data start row, và data end row
        
        Returns:
            (header_row, data_start_row, data_end_row)
        """
        header_row = None
        data_start_row = None
        data_end_row = None
        
        # Find header row
        for row in range(1, min(20, ws.max_row + 1)):
            first_cell = str(ws.cell(row, 1).value or '').lower()
            second_cell = str(ws.cell(row, 2).value or '').lower()
            third_cell = str(ws.cell(row, 3).value or '').lower()
            
            # Look for typical header keywords
            if any(kw in first_cell for kw in ['stt', 'no.', 'số tt']):
                header_row = row
                data_start_row = row + 1
                break
            
            # Alternative: check for "mã nhân viên" or "id"
            if any(kw in second_cell or kw in third_cell for kw in ['mã nv', 'mã nhân viên', 'employee id', 'id']):
                header_row = row
                data_start_row = row + 1
                break
        
        if not data_start_row:
            return None, None, None
        
        # Find data end row (before summary or empty rows)
        for row in range(data_start_row, min(ws.max_row + 1, data_start_row + 200)):
            first_cell = str(ws.cell(row, 1).value or '').lower()
            
            # Check if it's a summary row
            if any(kw in first_cell for kw in ['tổng', 'total', 'sum', 'subtotal']):
                data_end_row = row - 1
                break
            
            # Check if row is empty (no data in first 5 columns)
            has_data = any(ws.cell(row, col).value for col in range(1, 6))
            if not has_data:
                data_end_row = row - 1
                break
        
        if not data_end_row or data_end_row < data_start_row:
            data_end_row = data_start_row + 5  # Default
        
        return header_row, data_start_row, data_end_row
    
    def _identify_required_columns(self, ws, header_row: int) -> List[Tuple[int, str]]:
        """
        Identify required columns that should not be empty
        
        Returns:
            List of (col_idx, col_name) tuples
        """
        required_keywords = [
            "mã nhân viên", "mã nv", "employee id", "id",
            "họ và tên", "họ tên", "name",
            "lương cơ bản", "lương cb", "basic salary",
            "số ngày công", "ngày công", "working days"
        ]
        
        required_columns = []
        
        for col_idx in range(1, min(ws.max_column + 1, 100)):
            header = str(ws.cell(header_row, col_idx).value or '').lower().strip()
            
            if any(kw in header for kw in required_keywords):
                required_columns.append((col_idx, header))
        
        return required_columns
    
    def _check_empty_required_cells(
        self,
        ws,
        required_columns: List[Tuple[int, str]],
        data_start_row: int,
        data_end_row: int
    ) -> List[Dict[str, Any]]:
        """
        Check for empty cells in required columns
        
        Returns:
            List of warning dicts
        """
        warnings = []
        
        for col_idx, col_name in required_columns:
            empty_rows = []
            
            for row in range(data_start_row, data_end_row + 1):
                cell_value = ws.cell(row, col_idx).value
                
                # Check if empty (None, empty string, or formula resulting in empty)
                if cell_value is None or (isinstance(cell_value, str) and cell_value.strip() == ''):
                    empty_rows.append(row)
                
                # Check for error values
                if isinstance(cell_value, str) and cell_value.startswith('#'):
                    warnings.append({
                        "type": "CELL_ERROR",
                        "severity": "HIGH",
                        "message": f"Ô {get_column_letter(col_idx)}{row} trong cột '{col_name}' có lỗi: {cell_value}",
                        "cell": f"{get_column_letter(col_idx)}{row}",
                        "column": col_name,
                        "error_value": cell_value
                    })
            
            if empty_rows:
                # Only warn if many cells are empty
                if len(empty_rows) > 1:
                    warnings.append({
                        "type": "EMPTY_REQUIRED_CELLS",
                        "severity": "MEDIUM",
                        "message": f"Cột '{col_name}' có {len(empty_rows)} ô trống tại các dòng: {empty_rows[:5]}{'...' if len(empty_rows) > 5 else ''}",
                        "column": col_name,
                        "column_index": col_idx,
                        "empty_rows": empty_rows,
                        "count": len(empty_rows)
                    })
        
        return warnings
    
    def _check_formula_adjustment(
        self,
        ws,
        data_start_row: int,
        data_end_row: int
    ) -> List[Dict[str, Any]]:
        """
        Check if formulas are properly adjusted after row insertion
        
        Specifically:
        - Formulas should reference current row
        - Example: If row is 13, formula should be =B13*2, not =B12*2
        """
        issues = []
        
        for row in range(data_start_row, data_end_row + 1):
            for col_idx in range(1, min(ws.max_column + 1, 100)):
                cell = ws.cell(row, col_idx)
                
                if not isinstance(cell.value, str) or not cell.value.startswith('='):
                    continue
                
                formula = cell.value
                
                # Check for row references in formula
                # Pattern: A12, B13, etc.
                pattern = r'([A-Z]+)(\d+)'
                matches = re.findall(pattern, formula)
                
                if not matches:
                    continue
                
                # Check if any reference is using wrong row number
                wrong_row_refs = []
                for col_letter, row_num in matches:
                    row_num = int(row_num)
                    
                    # Skip if reference is to a different sheet or absolute reference
                    if "!" in formula:
                        continue
                    
                    # Check if row reference is significantly different from current row
                    # Allow +/- 1 for relative references, but flag larger differences
                    if abs(row_num - row) > 1 and row_num != 1:  # row 1 is likely header
                        wrong_row_refs.append((col_letter, row_num))
                
                if wrong_row_refs:
                    issues.append({
                        "type": "WRONG_ROW_REFERENCE",
                        "severity": "MEDIUM",
                        "message": f"Công thức tại {get_column_letter(col_idx)}{row} có thể tham chiếu sai dòng",
                        "cell": f"{get_column_letter(col_idx)}{row}",
                        "formula": formula,
                        "current_row": row,
                        "wrong_references": wrong_row_refs,
                        "details": f"Công thức tham chiếu tới dòng {wrong_row_refs[0][1]} nhưng đang ở dòng {row}"
                    })
        
        return issues
    
    def _check_vlookup_formulas(
        self,
        ws,
        data_start_row: int,
        data_end_row: int
    ) -> List[Dict[str, Any]]:
        """
        Check VLOOKUP and cross-sheet references
        
        Ensures:
        - VLOOKUP key references current row
        - Sheet references are valid
        """
        warnings = []
        
        for row in range(data_start_row, data_end_row + 1):
            for col_idx in range(1, min(ws.max_column + 1, 100)):
                cell = ws.cell(row, col_idx)
                
                if not isinstance(cell.value, str) or 'VLOOKUP' not in cell.value.upper():
                    continue
                
                formula = cell.value
                
                # Pattern: VLOOKUP(B13,'1.BCC'!C:D,2,FALSE)
                # Extract the lookup key (first parameter)
                vlookup_pattern = r'VLOOKUP\s*\(\s*([^,]+)'
                match = re.search(vlookup_pattern, formula, re.IGNORECASE)
                
                if not match:
                    continue
                
                lookup_key = match.group(1).strip()
                
                # Check if lookup key references current row
                key_row_match = re.search(r'([A-Z]+)(\d+)', lookup_key)
                if key_row_match:
                    key_col, key_row = key_row_match.groups()
                    key_row = int(key_row)
                    
                    if key_row != row:
                        warnings.append({
                            "type": "VLOOKUP_WRONG_ROW",
                            "severity": "HIGH",
                            "message": f"VLOOKUP tại {get_column_letter(col_idx)}{row} tham chiếu key từ dòng {key_row} thay vì dòng {row}",
                            "cell": f"{get_column_letter(col_idx)}{row}",
                            "formula": formula,
                            "current_row": row,
                            "key_row": key_row,
                            "suggestion": f"Đổi {key_col}{key_row} thành {key_col}{row}"
                        })
                
                # Check for sheet references
                sheet_ref_pattern = r"'([^']+)'!"
                sheet_refs = re.findall(sheet_ref_pattern, formula)
                
                for ref_sheet in sheet_refs:
                    if ref_sheet not in self.wb.sheetnames:
                        warnings.append({
                            "type": "INVALID_SHEET_REFERENCE",
                            "severity": "HIGH",
                            "message": f"VLOOKUP tại {get_column_letter(col_idx)}{row} tham chiếu sheet không tồn tại: '{ref_sheet}'",
                            "cell": f"{get_column_letter(col_idx)}{row}",
                            "formula": formula,
                            "missing_sheet": ref_sheet
                        })
        
        return warnings
    
    def _check_attendance_data(
        self,
        ws,
        header_row: int,
        data_start_row: int,
        data_end_row: int,
        keywords: List[str]
    ) -> List[Dict[str, Any]]:
        """Check if attendance columns have data"""
        warnings = []
        
        # Find attendance columns
        attendance_cols = []
        for col_idx in range(1, min(ws.max_column + 1, 110)):
            header = str(ws.cell(header_row, col_idx).value or '').lower()
            if any(kw in header for kw in keywords):
                attendance_cols.append((col_idx, header))
        
        if not attendance_cols:
            warnings.append({
                "type": "NO_ATTENDANCE_COLUMNS",
                "message": "Không tìm thấy cột chấm công (ngày công, số ngày làm việc, etc.)"
            })
            return warnings
        
        # Check each attendance column for data
        for col_idx, col_name in attendance_cols:
            empty_count = 0
            
            for row in range(data_start_row, data_end_row + 1):
                value = ws.cell(row, col_idx).value
                if value is None or (isinstance(value, str) and value.strip() == ''):
                    empty_count += 1
            
            if empty_count > (data_end_row - data_start_row + 1) / 2:
                warnings.append({
                    "type": "MOSTLY_EMPTY_ATTENDANCE_COLUMN",
                    "message": f"Cột chấm công '{col_name}' có {empty_count} ô trống trong tổng số {data_end_row - data_start_row + 1} dòng",
                    "column": col_name,
                    "empty_count": empty_count,
                    "total_rows": data_end_row - data_start_row + 1
                })
        
        return warnings
    
    def _find_ref_errors(self, ws) -> List[Dict[str, Any]]:
        """Find #REF! errors in sheet"""
        errors = []
        
        for row in range(1, min(ws.max_row + 1, 500)):
            for col in range(1, min(ws.max_column + 1, 100)):
                cell_value = ws.cell(row, col).value
                
                if isinstance(cell_value, str) and '#REF!' in cell_value:
                    errors.append({
                        "type": "REF_ERROR",
                        "severity": "HIGH",
                        "message": f"Lỗi #REF! tại {get_column_letter(col)}{row}",
                        "cell": f"{get_column_letter(col)}{row}",
                        "value": cell_value
                    })
        
        return errors
    
    def print_report(self, results: Dict[str, Any]):
        """Print validation report in Vietnamese"""
        print("\n" + "="*80)
        print(f"📋 BÁO CÁO VALIDATION KẾT QUẢ: {results['file']}")
        print("="*80)
        
        summary = results["summary"]
        
        # Overall status
        if results["success"]:
            print("\n✅ VALIDATION PASSED - Không có lỗi nghiêm trọng")
        else:
            print("\n❌ VALIDATION FAILED - Phát hiện lỗi cần xử lý")
        
        print(f"\nTổng quan:")
        print(f"  - Số lỗi (errors): {summary['total_errors']}")
        print(f"  - Số cảnh báo (warnings): {summary['total_warnings']}")
        print(f"  - Thông tin (info): {summary['total_info']}")
        print(f"  - Số sheet kiểm tra: {summary['sheets_validated']}")
        
        # Errors
        if results["errors"]:
            print(f"\n🔴 LỖI NGHIÊM TRỌNG ({len(results['errors'])}):")
            for i, error in enumerate(results["errors"], 1):
                print(f"\n  #{i}. [{error['type']}]")
                print(f"      {error['message']}")
                if 'cell' in error:
                    print(f"      Ô: {error['cell']}")
                if 'formula' in error:
                    print(f"      Công thức: {error['formula']}")
                if 'suggestion' in error:
                    print(f"      💡 Đề xuất: {error['suggestion']}")
        
        # Warnings
        if results["warnings"]:
            print(f"\n⚠️  CẢNH BÁO ({len(results['warnings'])}):")
            # Group by type
            warnings_by_type = {}
            for warning in results["warnings"]:
                wtype = warning['type']
                if wtype not in warnings_by_type:
                    warnings_by_type[wtype] = []
                warnings_by_type[wtype].append(warning)
            
            for wtype, warnings in warnings_by_type.items():
                print(f"\n  📌 {wtype} ({len(warnings)} trường hợp):")
                for warning in warnings[:3]:  # Show first 3
                    print(f"     - {warning['message']}")
                if len(warnings) > 3:
                    print(f"     ... và {len(warnings) - 3} trường hợp khác")
        
        # Info
        if results["info"]:
            print(f"\n💡 THÔNG TIN BỔ SUNG:")
            for info in results["info"]:
                print(f"  - {info['message']}")
        
        print("\n" + "="*80)
        print()


def validate_file(file_path: str, expected_employee_count: int = None):
    """Convenience function to validate a file and print report"""
    validator = ResultValidator(file_path)
    results = validator.validate_all(expected_employee_count)
    validator.print_report(results)
    return results


if __name__ == "__main__":
    import sys
    
    if len(sys.argv) < 2:
        print("Usage: python result_validator.py <excel_file> [expected_employee_count]")
        sys.exit(1)
    
    file_path = sys.argv[1]
    expected_count = int(sys.argv[2]) if len(sys.argv) > 2 else None
    
    validate_file(file_path, expected_count)
