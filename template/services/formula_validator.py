#!/usr/bin/env python3
"""
Validation tool to check for formula range issues in Excel files
"""
import openpyxl
import re
from typing import Dict, List, Tuple
import logging

# Logging configured in main.py
logger = logging.getLogger(__name__)


class FormulaValidator:
    """Validate Excel formulas for common issues"""
    
    def __init__(self, file_path: str):
        self.file_path = file_path
        self.wb = openpyxl.load_workbook(file_path, data_only=False)
        self.issues = []
    
    def validate_all_sheets(self) -> Dict:
        """Validate all sheets in workbook"""
        results = {
            "file": self.file_path,
            "sheets": [],
            "total_issues": 0
        }
        
        for sheet_name in self.wb.sheetnames:
            ws = self.wb[sheet_name]
            sheet_result = self.validate_sheet(ws, sheet_name)
            results["sheets"].append(sheet_result)
            results["total_issues"] += len(sheet_result["issues"])
        
        return results
    
    def validate_sheet(self, ws, sheet_name: str) -> Dict:
        """Validate formulas in a single sheet"""
        logger.info(f"Validating sheet: {sheet_name}")
        
        result = {
            "sheet_name": sheet_name,
            "issues": [],
            "data_start_row": None,
            "data_end_row": None,
            "summary_rows": []
        }
        
        # Find data range
        data_start, data_end = self._find_data_range(ws)
        result["data_start_row"] = data_start
        result["data_end_row"] = data_end
        
        if not data_start:
            return result
        
        # Find summary rows
        summary_rows = self._find_summary_rows(ws, data_start, data_end)
        result["summary_rows"] = summary_rows
        
        # Check formulas in summary rows
        for summary_row in summary_rows:
            issues = self._validate_summary_row(ws, summary_row, data_start, data_end)
            result["issues"].extend(issues)
        
        return result
    
    def _find_data_range(self, ws) -> Tuple[int, int]:
        """Find first and last rows with employee data"""
        data_start = None
        data_end = None
        
        # Look for header row first
        for row in range(1, min(20, ws.max_row + 1)):
            first_cell = str(ws.cell(row, 1).value or '').lower()
            if 'stt' in first_cell or 'mã nv' in first_cell:
                data_start = row + 1
                break
        
        if not data_start:
            # Try to find by checking for formulas
            for row in range(10, min(20, ws.max_row + 1)):
                has_data = any(ws.cell(row, col).value for col in range(1, 10))
                if has_data:
                    data_start = row
                    break
        
        if not data_start:
            return None, None
        
        # Find last data row before summary
        for row in range(data_start, min(ws.max_row + 1, data_start + 100)):
            first_cell = str(ws.cell(row, 1).value or '').lower()
            # Check if it's a summary row
            if 'tổng' in first_cell or 'total' in first_cell:
                data_end = row - 1
                break
            # Check if row is empty
            has_data = any(ws.cell(row, col).value for col in range(1, 10))
            if not has_data:
                data_end = row - 1
                break
        
        if not data_end:
            data_end = data_start + 10  # Default
        
        return data_start, data_end
    
    def _find_summary_rows(self, ws, data_start: int, data_end: int) -> List[int]:
        """Find rows that contain summary/aggregate formulas"""
        summary_rows = []
        
        # Check rows near data_end
        for row in range(data_start, min(ws.max_row + 1, data_end + 10)):
            first_cell = ws.cell(row, 1).value
            
            # Check if row has SUBTOTAL or aggregate formulas
            if first_cell and isinstance(first_cell, str) and first_cell.startswith('='):
                if 'SUBTOTAL' in first_cell.upper():
                    summary_rows.append(row)
                    continue
            
            # Check other cells for aggregate formulas
            for col in range(1, min(30, ws.max_column + 1)):
                cell_value = ws.cell(row, col).value
                if cell_value and isinstance(cell_value, str) and cell_value.startswith('='):
                    if self._is_aggregate_formula(cell_value):
                        if row not in summary_rows:
                            summary_rows.append(row)
                        break
        
        return summary_rows
    
    def _is_aggregate_formula(self, formula: str) -> bool:
        """Check if formula is an aggregate function"""
        formula_upper = formula.upper()
        aggregate_funcs = ['SUM(', 'SUBTOTAL(', 'AVERAGE(', 'COUNT(', 'SUMIF(']
        return any(func in formula_upper for func in aggregate_funcs)
    
    def _validate_summary_row(self, ws, summary_row: int, 
                              data_start: int, data_end: int) -> List[Dict]:
        """Validate formulas in summary row"""
        issues = []
        
        for col in range(1, min(50, ws.max_column + 1)):
            cell = ws.cell(summary_row, col)
            if not cell.value or not isinstance(cell.value, str):
                continue
            
            if not cell.value.startswith('='):
                continue
            
            formula = cell.value
            
            # Check for SUM formulas with wrong range
            if 'SUM(' in formula.upper():
                issue = self._check_sum_range(
                    formula, cell.coordinate, summary_row, data_start, data_end
                )
                if issue:
                    issues.append(issue)
        
        return issues
    
    def _check_sum_range(self, formula: str, cell_coord: str, 
                        summary_row: int, data_start: int, data_end: int) -> Dict:
        """Check if SUM range is correct"""
        # Pattern: SUM(A10:A20)
        pattern = r'SUM\(([A-Z]+)(\d+):([A-Z]+)(\d+)\)'
        matches = list(re.finditer(pattern, formula, re.IGNORECASE))
        
        if not matches:
            return None
        
        for match in matches:
            start_col = match.group(1)
            start_row = int(match.group(2))
            end_col = match.group(3)
            end_row = int(match.group(4))
            
            # Check if it's a vertical range
            if start_col != end_col:
                continue
            
            range_size = end_row - start_row + 1
            expected_size = data_end - data_start + 1
            
            # Issue 1: Single row range (likely wrong)
            if range_size == 1 and start_row == summary_row:
                return {
                    "cell": cell_coord,
                    "formula": formula,
                    "issue_type": "SINGLE_ROW_SUM",
                    "severity": "HIGH",
                    "description": f"SUM formula only sums row {summary_row} (itself)",
                    "expected": f"Should sum from row {data_start} to {data_end}",
                    "suggested_fix": formula.replace(
                        f"{start_col}{start_row}:{end_col}{end_row}",
                        f"{start_col}{data_start}:{end_col}{data_end}"
                    )
                }
            
            # Issue 2: Range doesn't cover all data
            elif start_row > data_start or end_row < data_end:
                return {
                    "cell": cell_coord,
                    "formula": formula,
                    "issue_type": "INCOMPLETE_RANGE",
                    "severity": "MEDIUM",
                    "description": f"SUM range ({start_row}-{end_row}) doesn't cover all data ({data_start}-{data_end})",
                    "expected": f"Should sum from row {data_start} to {data_end}",
                    "suggested_fix": formula.replace(
                        f"{start_col}{start_row}:{end_col}{end_row}",
                        f"{start_col}{data_start}:{end_col}{data_end}"
                    )
                }
        
        return None
    
    def print_report(self, results: Dict):
        """Print validation report"""
        print("\n" + "="*80)
        print(f"📋 FORMULA VALIDATION REPORT: {results['file']}")
        print("="*80)
        
        total_issues = results['total_issues']
        if total_issues == 0:
            print("\n✅ No issues found! All formulas are correct.\n")
            return
        
        print(f"\n⚠️  Found {total_issues} issue(s)\n")
        
        for sheet in results['sheets']:
            if not sheet['issues']:
                continue
            
            print(f"\n📄 Sheet: {sheet['sheet_name']}")
            print(f"   Data range: Row {sheet['data_start_row']} to {sheet['data_end_row']}")
            print(f"   Summary rows: {sheet['summary_rows']}")
            print(f"   Issues: {len(sheet['issues'])}")
            
            for i, issue in enumerate(sheet['issues'], 1):
                severity_icon = "🔴" if issue['severity'] == 'HIGH' else "🟡"
                print(f"\n   {severity_icon} Issue #{i}: {issue['issue_type']}")
                print(f"      Cell: {issue['cell']}")
                print(f"      Formula: {issue['formula']}")
                print(f"      Problem: {issue['description']}")
                print(f"      Expected: {issue['expected']}")
                print(f"      Suggested fix: {issue['suggested_fix']}")


def validate_file(file_path: str):
    """Validate a single file"""
    validator = FormulaValidator(file_path)
    results = validator.validate_all_sheets()
    validator.print_report(results)
    return results


if __name__ == "__main__":
    import sys
    
    # Validate output files
    files = [
        "uploads/output_final.xlsx",
        "uploads/output_test.xlsx", 
        "uploads/output_verified.xlsx"
    ]
    
    for file_path in files:
        try:
            validate_file(file_path)
        except Exception as e:
            print(f"Error validating {file_path}: {e}")
